#!/usr/bin/env python3
#-*- coding: utf-8 -*-
# -*- mode: python; python-indent-offset: 4 -*-
#
# ./copy_colored_text.py filename
#
# Purpose: Create entries in columns for "eng-draft" and "sv-fraft"
#          based upon those entries in the 'eng' and 'sv' columns that were
#          set in red (specifically rgb:  #FFFF0000') as these represent
#          my proposed words.
#
#	   The draft columns are to facilitate a native Swedish speaker
#	   evaluating the proposed 'eng' or 'sv' entries set in red.
#
# Summary: The program opens the file and copies ext from "eng" to "eng-draft"
#          if 'eng' text is red and copies text from "sv" to "sv-draft" if text is red
#
#
# Input: This program takes a spreadsheet of data based upon those found
#        in Computer Sweden's "IT-ord: Ord och uttryck i it-branschen"
# 	 See https://it-ord.idg.se/
#
# The wordlist is a collection of words that are related to IT.
# Here "words" are strings that can include spaces and commas.
#
# As part of the STUNDA activities, I have attempted to find
# English and Swedish word pairs. When either language was missing,
# I have tried with the help of Google's Gemini to identify a suitable word or words. 
#
# I have also tried to classify the words into categories or combinations of
# categories (for examaple, a given word might be the name of a product and
# a company name).
#
# Outut: an updated spread sheet in temp.xlsx 
#
# ----------------------------------------------------------------------
# SPREADSHEET
#
# The spreadshet has the following columns:
# Column A:  		contains a number (starting from 0) indicating this was
#                       the n-th word that I extracted from the word list.
#                       Rows that were based on words found when examining
#                       the web page for a particular term, that did not exist
#                       in the list of words - have an empty column A
#
# Column B: entry	the "word" from the wordlist
#
# Column C: PI    	Indicates (Tue or False) if the original word ended with an arrow
#
# Column D: URL  	contains a URL to the source page
#                       The  URL that points to the definition for this word (term)
#                       in Computer Sweden's wordlist.
#                       Thus the interested reader can see the full source.
#
# Column E: 'short for or acronym' 	contains acronyms, abbreviations, or short forms
#
# Column F: eng   	contains the English form of the word
#
# Column G: sv		contains the Swedish form of the word
#
# Column H: eng-draft	contains a potential English version 
#                       as reviewed by a native Swedish speaker
#
# Column I: sv-draft	contains a potential Swedish version
#                       as reviewed by a native Swedish speaker
#
# Column J: Swedish Description	contains
#			 a Swedish descritpion of the word - generally based
#                        on the words' entry in the word list
#                        [This text is NOT intended to be used beyond the purposes
#                        of correct the "eng-draft" and "sv-fraft" entries.
#                        This column should be removed before providing
#                        the spreadsheet to parties outside those working on STUNDA.]
#
# Column K: category	contains one or more categories that I have tried to define
#
# Column L: English comments	contains my own collments in English
#
# ----------------------------------------------------------------------
# Conventions:
# Strings set in red (specifically rgb:  #FFFF0000') are strings that
# I have added and do NOT represent a word in the IT-ord entries
#
# When there are mupltiple words separated by commas, these are:
#   A. a comma separared phrase (in which case there is a conjunciton before the final word or
#   B. a list of words that are all candidates for the relevant entry
#
# When there are mupltiple words separated by semicolons, these are either separate words or separate words with different meanings (corresponding to different definitions given in the IT-ord entry for the word.
#
# Some of the categories are based upon the categories used in the web pages
# of the IT-ord wordlist (i.e., 'nät- och sms-språk' and jargon), while the other
# categories are my attempt to categorize the entry.
# The assignment to categories is loose and some categories should perhaps be merged,
# such as "court case" and "legal case".
#
# One of the improant categories is "metaentry", which indicated a word for which
# there are multiple (logical) different entries - with different meanings.
# These entries in this category should be considered metadata and not actual "words".
#
# Examples:
# ./copy_colored_text.py cs_words_ss-edited-20250314.xlsx
#
# Notes
# Extensive use is made of the openpyxl.
# Due to the use of red for _part_ of a string, it is important to understand the
# concept of Rich Text. In the underlying XML Rich Text enables a cell in the
# spreadsheet to contain CellRichText - which is logically a list of either
# Text or TextBlock (or a mix of these). Each TextBlock can contain formatting
# information (using InlineFont) and a string.
#   For details see: https://openpyxl.readthedocs.io/en/stable/rich_text.html
#
# The choice of the second red like color (#FFFF0000) was not a careful choice,
# but was simply chosen because:
#  (a) I could distinguish it from the default (black) text and
#  (b) it was easy to click on!
#
# Acknowledgements
# I would like to acknowledge the hlp of Google Gemini Advanced (2.0 Flash)
#  (a) for help finding missing English and Swedish terms.
#      The prompts that were used were generaall of the form:
#       Computer Sweden defines xxxxx
#                 part of the definition from the IT-ord web page
#       What is the English/Swedish term for yyyyy?
#
#      If xxxxx was an acronym or short form, then yyyyy was the spelled out version.
#
# (b) for help with processing CellRichText
#      While the suggested python code was incorrect and would not actually
#      produce the desired output, it suggested directions for me to look.
#      See the test program check_color.py
#
#
# 2025-03-14
#
# G. Q. Maguire Jr.
#
import csv, time
from pprint import pprint
import optparse
import sys


from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.cell.rich_text import Text
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

def check_red_substring(cell):
    """Checks if any substring within a cell is red (RGB: FFFF0000)."""
    if cell.value is None:
        return False

    if cell.font and cell.font.color and cell.font.color.rgb == 'FFFF0000':
        return True # if the entire cell is red

    if type(cell.value) == CellRichText:
        for element in cell.value:
            if isinstance(element, Text):
                if element.font and element.font.color and element.font.color.rgb == 'FFFF0000':
                    return True
            if isinstance(element, TextBlock):
                if element.font and element.font.color and element.font.color.rgb == 'FFFF0000':
                    return True

    return False



def get_columns_from_worksheet(ws):
  return {
      cell.value: {
          'letter': get_column_letter(cell.column),
          'number': cell.column - 1
      } for cell in ws[1] if cell.value
  }


def main():
    global Verbose_Flag
    #
    parser = optparse.OptionParser()
    #
    parser.add_option('-v', '--verbose',
                      dest="verbose",
                      default=False,
                      action="store_true",
                      help="Print lots of output to stdout"
    )
    #
    options, remainder = parser.parse_args()
    #
    Verbose_Flag=options.verbose
    if Verbose_Flag:
        print("ARGV      : {}".format(sys.argv[1:]))
        print("VERBOSE   : {}".format(options.verbose))
        print("REMAINING : {}".format(remainder))
    #
    if (len(remainder) < 1):
        print("Insuffient arguments\n must provide file_name\n")
        return
    #
    input_file_name=remainder[0]
    wb = load_workbook(filename = input_file_name, rich_text=True)
    ws = wb.active
    print(f'{ws.title=}')

    row_count = ws.max_row
    column_count = ws.max_column
    print(f'{row_count=} {column_count=}')


    rows = ws.iter_rows(min_row=1, max_row=1) # returns a generator of rows
    first_row = next(rows) # get the first row
    headings = [c.value for c in first_row] # extract the values from the cells
    print(f'{headings=}')

    COLUMNS = get_columns_from_worksheet(ws)
    print(f'{COLUMNS=}')

    eng_column=COLUMNS['eng']['number']+1
    sv_column=COLUMNS['sv']['number']+1
    eng_draft_column=COLUMNS['eng-draft']['number']+1
    sv_draft_column=COLUMNS['sv-draft']['number']+1
    category_column=COLUMNS['category']['number']+1


    print(f'{eng_column=} {sv_column=} {eng_draft_column=} {sv_draft_column=}')
    # process English columns
    for index, row in enumerate(ws.rows, start=2):
        c = ws.cell(index, eng_column)
        #print(f'{index=} {c=}  {c.value=}  {c.font=}')
        if check_red_substring(c):
            if Verbose_Flag:
                print(f'{index=} {c=} {c.value=} {c.font.color.rgb=}')
            ws.cell(index, eng_draft_column).value=c.value
            if type(c.value) != CellRichText:
                if c.font and c.font.color and c.font.color.rgb == 'FFFF0000':
                    ws.cell(index, eng_draft_column).font=Font(color='FFFF0000')

    # process Swedish columns
    for index, row in enumerate(ws.rows, start=2):
        c = ws.cell(index, sv_column)
        if Verbose_Flag:
            print(f'Cell: {index=} {c=}  {c.value=}  {c.font=}')
        if check_red_substring(c):
            if Verbose_Flag:
                print(f'Red substring: {index=} {c=} {c.value=} {c.font.color.rgb=}')
            ws.cell(index, sv_draft_column).value=c.value
            if type(c.value) != CellRichText:
                if c.font and c.font.color and c.font.color.rgb == 'FFFF0000':
                    ws.cell(index, sv_draft_column).font=Font(color='FFFF0000')

    categories_set=set()
    categories_stats=dict()
    original_categories=set()
    
    # process category column
    for index, row in enumerate(ws.rows, start=2):
        c = ws.cell(index, category_column)
        if c.value and c.value[0] != '=': # skip formulas
            w=c.value.strip()
            original_categories.add(w)
            if ';' in w:
                wrds=w.split(';')
                for nw in wrds:
                    s1=nw.strip()
                    categories_set.add(s1)
                    categories_stats[s1]=categories_stats.get(s1, 0) + 1

            elif ',' in w:
                wrds=w.split(',')
                for nw in wrds:
                    s1=nw.strip()
                    categories_set.add(s1)
                    categories_stats[s1]=categories_stats.get(s1, 0) + 1
            else:
                categories_set.add(w)
                categories_stats[w]=categories_stats.get(w, 0) + 1
                
    categories=sorted(categories_set)
    original_categories=sorted(original_categories)
    print(f'{len(original_categories)=} {original_categories=}')

    print(f'{len(categories_set)=} {categories=}')

    first_row="{COLUMNS['category']['letter']}3"

    index_offset= row_count + 9
    ws.cell(index_offset, category_column+1).value=CellRichText(
        TextBlock(InlineFont(sz=24, b=True, color=Color(rgb='FFFF0000')), "Original categories")
    )

    
    last_index_offset=0
    for index, wrds in enumerate(original_categories):
        index_offset=index + row_count + 10
        last_index_offset=index_offset
        ws.cell(index_offset, category_column+1).value=wrds
        ws.cell(index_offset, category_column).value=f"=COUNTIF({COLUMNS['category']['letter']}3:{COLUMNS['category']['letter']}{row_count}, {COLUMNS['English comments']['letter']}{index_offset})"

    
    index_offset= last_index_offset + 9
    ws.cell(index_offset, category_column+1).value=CellRichText(
        TextBlock(InlineFont(sz=24, b=True, color=Color(rgb='FFFF0000')), "Categories")
    )


    for index, wrds in enumerate(categories):
        index_offset=last_index_offset + index + 10
        ws.cell(index_offset, category_column+1).value=wrds
        ws.cell(index_offset, category_column).value=categories_stats.get(wrds, '')




    wb.save('temp.xlsx')
#
if __name__ == "__main__": main()

