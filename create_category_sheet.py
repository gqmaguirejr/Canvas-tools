#!/usr/bin/env python3
#-*- coding: utf-8 -*-
# -*- mode: python; python-indent-offset: 4 -*-
#
# ./create_category_sheet.py filename [category]+
#
# Purpose: Create a new sheet in spread sheet for the specified caterory or categories
#
#
# Input: This program takes a spreadsheet of data based upon those found
#        in Computer Sweden's "IT-ord: Ord och uttryck i it-branschen"
# 	 See https://it-ord.idg.se/
#
# The wordlist is a collection of words that are related to IT.
# Here "words" are strings that can include spaces and commas.
#
# As part of the STUNDA activities, I have attempted to find
# English and Swedish word pairs. When either language was missing,
# I have tried with the help of Google's Gemini to identify a suitable word or words. 
#
# I have also tried to classify the words into categories or combinations of
# categories (for examaple, a given word might be the name of a product and
# a company name).
#
# Outut: an updated spread sheet in temp.xlsx 
#
# ----------------------------------------------------------------------
# SPREADSHEET
#
# The spreadshet has the following columns:
# Column A:  		contains a number (starting from 0) indicating this was
#                       the n-th word that I extracted from the word list.
#                       Rows that were based on words found when examining
#                       the web page for a particular term, that did not exist
#                       in the list of words - have an empty column A
#
# Column B: entry	the "word" from the wordlist
#
# Column C: PI    	Indicates (Tue or False) if the original word ended with an arrow
#
# Column D: URL  	contains a URL to the source page
#                       The  URL that points to the definition for this word (term)
#                       in Computer Sweden's wordlist.
#                       Thus the interested reader can see the full source.
#
# Column E: 'short for or acronym' 	contains acronyms, abbreviations, or short forms
#
# Column F: eng   	contains the English form of the word
#
# Column G: sv		contains the Swedish form of the word
#
# Column H: eng-draft	contains a potential English version 
#                       as reviewed by a native Swedish speaker
#
# Column I: sv-draft	contains a potential Swedish version
#                       as reviewed by a native Swedish speaker
#
# Column J: Swedish Description	contains
#			 a Swedish descritpion of the word - generally based
#                        on the words' entry in the word list
#                        [This text is NOT intended to be used beyond the purposes
#                        of correct the "eng-draft" and "sv-fraft" entries.
#                        This column should be removed before providing
#                        the spreadsheet to parties outside those working on STUNDA.]
#
# Column K: category	contains one or more categories that I have tried to define
#
# Column L: English comments	contains my own collments in English
#
# ----------------------------------------------------------------------
# Conventions:
# Strings set in red (specifically rgb:  #FFFF0000') are strings that
# I have added and do NOT represent a word in the IT-ord entries
#
# When there are mupltiple words separated by commas, these are:
#   A. a comma separared phrase (in which case there is a conjunciton before the final word or
#   B. a list of words that are all candidates for the relevant entry
#
# When there are mupltiple words separated by semicolons, these are either separate words or separate words with different meanings (corresponding to different definitions given in the IT-ord entry for the word.
#
# Some of the categories are based upon the categories used in the web pages
# of the IT-ord wordlist (i.e., 'nät- och sms-språk' and jargon), while the other
# categories are my attempt to categorize the entry.
# The assignment to categories is loose and some categories should perhaps be merged,
# such as "court case" and "legal case".
#
# One of the improant categories is "metaentry", which indicated a word for which
# there are multiple (logical) different entries - with different meanings.
# These entries in this category should be considered metadata and not actual "words".
#
# Examples:
# ./create_category_sheet.py /z3/maguire/STUNDA/Computer_Sweden/cs_words_ss-edited-20250322.xlsx 'company'
#
# Notes
# Extensive use is made of the openpyxl.
# Due to the use of red for _part_ of a string, it is important to understand the
# concept of Rich Text. In the underlying XML Rich Text enables a cell in the
# spreadsheet to contain CellRichText - which is logically a list of either
# Text or TextBlock (or a mix of these). Each TextBlock can contain formatting
# information (using InlineFont) and a string.
#   For details see: https://openpyxl.readthedocs.io/en/stable/rich_text.html
#
# The choice of the second red like color (#FFFF0000) was not a careful choice,
# but was simply chosen because:
#  (a) I could distinguish it from the default (black) text and
#  (b) it was easy to click on!
#
# Acknowledgements
# I would like to acknowledge the hlp of Google Gemini Advanced (2.0 Flash)
#  (a) for help finding missing English and Swedish terms.
#      The prompts that were used were generaall of the form:
#       Computer Sweden defines xxxxx
#                 part of the definition from the IT-ord web page
#       What is the English/Swedish term for yyyyy?
#
#      If xxxxx was an acronym or short form, then yyyyy was the spelled out version.
#
# (b) for help with processing CellRichText
#      While the suggested python code was incorrect and would not actually
#      produce the desired output, it suggested directions for me to look.
#      See the test program check_color.py
#
#
# 2025-03-14
#
# G. Q. Maguire Jr.
#
import csv, time
from pprint import pprint
import optparse
import sys

sys.path.append('/z3/maguire/Canvas/Canvas-tools')  # Include the path to module_folder
#  as common_English_words, common_swedish_words, common_swedish_technical_words
import common_english
import common_swedish
import common_acronyms
import common_strings
import AVL_words_with_CEFR

# width to use for outputting numeric values
Numeric_field_width=7


from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import colors
from openpyxl.styles import Font, Color
from openpyxl.cell.rich_text import Text
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

def check_red_substring(cell):
    """Checks if any substring within a cell is red (RGB: FFFF0000)."""
    if cell.value is None:
        return False

    if cell.font and cell.font.color and cell.font.color.rgb == 'FFFF0000':
        return True # if the entire cell is red

    if type(cell.value) == CellRichText:
        for element in cell.value:
            if isinstance(element, Text):
                if element.font and element.font.color and element.font.color.rgb == 'FFFF0000':
                    return True
            if isinstance(element, TextBlock):
                if element.font and element.font.color and element.font.color.rgb == 'FFFF0000':
                    return True

    return False



def get_columns_from_worksheet(ws):
  return {
      cell.value: {
          'letter': get_column_letter(cell.column),
          'number': cell.column - 1
      } for cell in ws[1] if cell.value
  }


IT_ord_persons=[
    'Abrams, Jenna', # person
    'Aiken, Howard', # person
    'Allen, Frances', # person
    'Allen, Paul', # person
    'Andreessen, Marc', # person
    'Antkare, Ike', # person
    'Aronsson, Lars', # person
    'Arraf, Amina', # person
    'Atanasoff, John Vincent', # person
    'Ayyadurai, Shiva', # person
    'Babbage, Charles', # person
    'Badawi, Raif', # person
    'Bara, Smaranda', # person
    'Baran, Paul', # person
    'Barlow, John Perry', # person
    'Benford, Gregory', # person
    'Berezin, Evelyn', # person
    'Berners-Lee, Tim', # person
    'Berry, Daniella Bunten', # person
    'Beurling, Arne', # person
    'Bézier', # person
    'Bina, Eric', # person
    'Bini, Ola', # person
    'Boggs, David', # person
    'Borg, Anita', # person
    'Brand, Stewart', # person
    'Bush, Vannevar', # person
    'Byron, Ada', # person
    'Cailliau, Robert', # person
    'Chatterjee, Satrajit', # person
    'Chaum, David', # person
    'Church, Alonzo', # person
    'Clark, Jim', # person
    'Codd, E F', # person
    'Cray, Seymour', # person
    'Cringely, Robert X', # person
    'Dahl, Ole-Johan', # person
    'Damore, James', # person
    'Diffie, Whitfield', # person
    'Dijkstra, Edsger', # person
    'Dubinsky, Donna', # person
    'Duthel, Heinz', # person
    'Dvorak, John', # person
    'Eckert, J Presper', # person
    'Engelbart, Douglas', # person
    'Estrin, Deborah', # person
    'Fawkes, Guy', # person
    'Feynman, Richard', # person
    'Flowers, Tom', # person
    'Freese, Jan', # person
    'Friis, Janus', # person
    'Gates, Bill', # person
    'Gebru, Timnit', # person
    'Gibson, William', # person
    'Gilmore, John', # person
    'Gödel, Kurt', # person
    'Hagström, Stig', # person
    'Hamilton, Margaret', # person
    'Hawes, Mary K', # person
    'Hawkins, Jeff', # person
    'Hellman, Martin', # person
    'Hertzfeld, Andy', # person
    'Hopper, Grace', # person
    'Icaza, de, Miguel', # person
    'Jacobson, Ivar', # person
    'Jobs, Steve', # person
    'Johansen, Jon Lech', # person
    'Jones, Alex', # person
    'Joy, Bill', # person
    'Kapor, Mitch', # person
    'Kare, Susan', # person
    'Karlqvist, Olle', # person
    'Kay, Alan', # person
    'Kilby, Jack', # person
    'Kildall, Dorothy',
    'McEwen, Dorothy', # person
    'Kildall, Gary', # person
    'Kleinberg, Jon', # person
    'Kleinrock, Leonard', # person
    'Knuth, Don', # person
    'Kurzweil, Ray', # person
    'Lanier, Jaron', # person
    'Lans, Håkan', # person
    'Lorenz, Edward', # person
    'Love, Lauri', # person
    'Lovelace, Ada', # person
    'Mandelbrot, Benoit', # person
    'Margaret Mitchell', # person
    'Mauchly, John', # person
    'McEwen, Dorothy', # person
    'McGrath, Titania', # person
    'McKinnon, Gary', # person
    'Metcalfe, Bob', # person
    'Moggridge, Bill', # person
    'Moore, Gordon', # person
    'Müller, Johann', # person
    'Musk, Elon', # person
    'Nakamoto, Satoshi', # person
    'Negroponte, Nicholas', # person
    'Neij, Fredrik', # person
    'Nelson, Ted', # person
    'Neumann, von, John', # person
    'Norris, Chuck', # person
    'Noyce, Robert', # person
    'Nunes, Devin', # person
    'Nygaard, Kristen', # person
    'Odlyzko, Andrew', # person
    'Olsen, Ken', # person
    'Olsson, Anders R', # person
    'Osborne, Adam', # person
    'Otlet, Paul', # person
    'Ozzie, Ray', # person
    'Parker, Philip M', # person
    'Patterson, David', # person
    'Perlman, Radia', # person
    'Polese, Kim', # person
    'Poole, Christopher', # person
    'Quinn, Zoë', # person
    'Raskin, Jef', # person
    'Raymond, Eric', # person
    'Rheingold, Howard', # person
    'Rhodes, Ida', # person
    'Ritchie, Dennis', # person
    'Roberts, Larry', # person
    'Ross, Thomas S', # person
    'Rubin, Andy', # person
    'Sammet, Jean', # person
    'Sarkeesian, Anita', # person
    'Schneier, Bruce', # person
    'Shannon, Claude', # person
    'Shirley, Stephanie', # person
    'Sholes, Christopher', # person
    'Snowden, Edward', # person
    'Stallman, Richard', # person
    'Stenhagen, Jan-Jöran', # person
    'Stephens, Mark', # person
    'Sterling, Bruce', # person
    'von Neumann, John', # person
    'Ware, Willis', # person
    'Weiser, Mark', # person
    'Wiberg, Martin', # person
    'Wiener, Norbert', # person
    'Wolfram, Stephen', # person
    'Wozniak, Steve', # person
    'Zachary, G Pascal', # person, author
    'Zennström, Niklas', # person, inventor and entrepreneur
    'Zimmermann, Phil', # person
    'Zuse, Konrad', # person
]


def main():
    global Verbose_Flag
    #
    parser = optparse.OptionParser()
    #
    parser.add_option('-v', '--verbose',
                      dest="verbose",
                      default=False,
                      action="store_true",
                      help="Print lots of output to stdout"
    )

    parser.add_option('-a', '--acronyms',
                      dest="acronym_option",
                      default=False,
                      action="store_true",
                      help="process acronyms"
    )
    #
    options, remainder = parser.parse_args()
    #
    Verbose_Flag=options.verbose
    if Verbose_Flag:
        print("ARGV      : {}".format(sys.argv[1:]))
        print("VERBOSE   : {}".format(options.verbose))
        print("REMAINING : {}".format(remainder))
    #
    if (len(remainder) < 1):
        print("Insuffient arguments\n must provide file_name\n")
        return
    #
    input_file_name=remainder[0]
    wb = load_workbook(filename = input_file_name, rich_text=True)
    ws = wb.active
    print(f'{ws.title=}')

    row_count = ws.max_row
    column_count = ws.max_column
    print(f'{row_count=} {column_count=}')


    rows = ws.iter_rows(min_row=1, max_row=1) # returns a generator of rows
    first_row = next(rows) # get the first row
    headings = [c.value for c in first_row] # extract the values from the cells
    print(f'{headings=}')

    COLUMNS = get_columns_from_worksheet(ws)
    print(f'{COLUMNS=}')

    eng_column=COLUMNS['eng']['number']+1
    sv_column=COLUMNS['sv']['number']+1
    eng_draft_column=COLUMNS['eng-draft']['number']+1
    sv_draft_column=COLUMNS['sv-draft']['number']+1
    category_column=COLUMNS['category']['number']+1
    acronym_column=COLUMNS['short for or acronym']['number']+1
    

    print(f'{eng_column=} {sv_column=} {eng_draft_column=} {sv_draft_column=}')

    categories_set=set()
    categories_stats=dict()
    original_categories=set()
    
    # process category column
    for index, row in enumerate(ws.rows, start=2):
        c = ws.cell(index, category_column)
        if isinstance(c.value, int):
            continue
        if type(c.value) != str:
            continue
        if c.value and c.value[0] != '=': # skip formulas
            w=c.value.strip()
            original_categories.add(w)
            if ';' in w:
                wrds=w.split(';')
                for nw in wrds:
                    s1=nw.strip()
                    categories_set.add(s1)
                    categories_stats[s1]=categories_stats.get(s1, 0) + 1

            elif ',' in w:
                wrds=w.split(',')
                for nw in wrds:
                    s1=nw.strip()
                    categories_set.add(s1)
                    categories_stats[s1]=categories_stats.get(s1, 0) + 1
            else:
                categories_set.add(w)
                categories_stats[w]=categories_stats.get(w, 0) + 1
                
    categories=sorted(categories_set)
    original_categories=sorted(original_categories)
    print(f'{len(original_categories)=} {original_categories=}')

    print(f'{len(categories_set)=} {categories=}')


    # get the catergory to filter on
    filter_categories=list()
    if (len(remainder) > 1):
        for i in range(1, len(remainder)):
            if Verbose_Flag:
                print(f"{i}: {remainder[i]}")
            if remainder[i] in categories:
                filter_categories.append(remainder[i])
            else:
                print(f'{remainder[i]=} not in available categories')

    print(f'{filter_categories=}')

    for fc in filter_categories:
        if fc not in wb.sheetnames:
            fc_sheet=wb.create_sheet(fc)
            #
            # copy first row of 'Words' sheet to fc_sheet
            for index, row in enumerate(ws.rows, start=0):
                fc_sheet.append((cell.value for cell in row))
                if index == 0:
                    break

            # copy rows of 'Words' sheet to fc_sheet if the category matches
            for index, row in enumerate(ws.rows, start=1):
                cell_categories_set=set()
                c = ws.cell(index, category_column)
                if c.value is None or isinstance(c.value, int):
                    continue

                if not isinstance(c.value, str):
                    print(f'{c.value=} is not a string')
                    continue

                if c.value and c.value[0] == '=': # skip formulas
                    continue

                w=c.value.strip()

                if ';' in w:
                    wrds=w.split(';')
                    for nw in wrds:
                        s1=nw.strip()
                        cell_categories_set.add(s1)
                elif ',' in w:
                    wrds=w.split(',')
                    for nw in wrds:
                        s1=nw.strip()
                        cell_categories_set.add(s1)
                else:
                    cell_categories_set.add(w)

                if fc in cell_categories_set:
                    #print(f'{index=} {row=}')
                    fc_sheet.append((cell.value for cell in row))



    # # process English columns
    # for index, row in enumerate(ws.rows, start=2):
    #     c = ws.cell(index, eng_column)
    #     #print(f'{index=} {c=}  {c.value=}  {c.font=}')
    #     if check_red_substring(c):
    #         if Verbose_Flag:
    #             print(f'{index=} {c=} {c.value=} {c.font.color.rgb=}')
    #         ws.cell(index, eng_draft_column).value=c.value
    #         if type(c.value) != CellRichText:
    #             if c.font and c.font.color and c.font.color.rgb == 'FFFF0000':
    #                 ws.cell(index, eng_draft_column).font=Font(color='FFFF0000')

    # # process Swedish columns
    # for index, row in enumerate(ws.rows, start=2):
    #     c = ws.cell(index, sv_column)
    #     if Verbose_Flag:
    #         print(f'Cell: {index=} {c=}  {c.value=}  {c.font=}')
    #     if check_red_substring(c):
    #         if Verbose_Flag:
    #             print(f'Red substring: {index=} {c=} {c.value=} {c.font.color.rgb=}')
    #         ws.cell(index, sv_draft_column).value=c.value
    #         if type(c.value) != CellRichText:
    #             if c.font and c.font.color and c.font.color.rgb == 'FFFF0000':
    #                 ws.cell(index, sv_draft_column).font=Font(color='FFFF0000')


    wb.save('temp.xlsx')

    unique_words=set()
    # entries in the dict will have the form: 'acronym': ['expanded form 1', 'expanded form 2',  ... ]
    well_known_acronyms=dict()
    for e in common_acronyms.well_known_acronyms_list:
        if len(e) >= 1:
            ack=e[0]
            if len(e) >= 2:
                d=e[1]
                current_entry=well_known_acronyms.get(ack, list())
                current_entry.append(d)
                well_known_acronyms[ack]=current_entry
    print(f'{(len(well_known_acronyms)):>{Numeric_field_width}} unique acronyms in ({len(common_acronyms.well_known_acronyms_list)}) well_known_acronyms')

    # process English columns - collecting words
    for fc in filter_categories:
        for index, row in enumerate(ws.rows, start=2):
            c = ws.cell(index, eng_column)

            cell_categories_set=set()
            ctg = ws.cell(index, category_column)
            if ctg.value is None or isinstance(ctg.value, int):
                continue

            if not isinstance(ctg.value, str):
                print(f'{ctg.value=} is not a string')
                continue

            if ctg.value and ctg.value[0] == '=': # skip formulas
                continue

            w=ctg.value.strip()
            if ';' in w:
                wrds=w.split(';')
                for nw in wrds:
                    s1=nw.strip()
                    cell_categories_set.add(s1)
            elif ',' in w:
                wrds=w.split(',')
                for nw in wrds:
                    s1=nw.strip()
                    cell_categories_set.add(s1)
            else:
                cell_categories_set.add(w.strip())

            if fc in cell_categories_set:
                unique_words.add(c.value)

    print(f'{len(unique_words)=}, {unique_words=}')

    print('New entries::::')

    # collect English & Swedish expansions of knonw acronyms
    known_E_acronyms=set()
    known_S_acronyms=set()
    for w in common_acronyms.well_known_acronyms_list:
        if len(w) < 2:
            continue
        if w[1]:
            if isinstance(w[1], dict):
                print(f"dict!!!! {w=}")
            elif isinstance(w[1], str):
                known_E_acronyms.add(w[1])
            else:
                print(f"unknown case!!!! {w=}")

        if len(w) > 2:
            if not isinstance(w[2], dict):
                continue
            known_S_acronyms.add(w[2].get('swe'))

    # now check to see of this word is known already
    new_words=[]
    for w in unique_words:

        if isinstance(w, float):
            print(f"float::: '{w}'")
            continue

        if isinstance(w, int):
            print(f"float::: '{w}'")
            continue


        if w in common_english.common_English_words:
            continue

        if w in common_english.company_and_product_names:
            continue

        if w in common_english.common_programming_languages:
            continue

        if w in common_english.place_names:
            continue

        if w in common_english.common_units:
            continue

        if w in common_english.common_french_words:
            continue

        if w in common_english.common_danish_words:
            continue
        if w in common_english.common_finnish_words:
            continue

        if w in common_english.common_german_words:
            continue

        if w in common_english.common_icelandic_words:
            continue

        if w in common_english.common_italian_words:
            continue

        if w in common_english.common_japanese_words:
            continue

        if w in common_english.common_latin_words:
            continue

        if w in common_english.common_norwegian_words:
            continue

        if w in common_english.common_portuguese_words:
            continue

        if w in common_english.common_spanish_words:
            continue

        if w in common_english.programming_keywords:
            continue

        if w in common_english.language_tags:
            continue

        if w in common_english.mathematical_words_to_ignore:
            continue

        if w in common_english.abbreviations_ending_in_period:
            continue

        if w in common_english.amino_acids:
            continue

        if w in common_english.binary_prefixes:
            continue

        if w in common_english.decimal_prefixes:
            continue

        if w in known_E_acronyms or w in known_S_acronyms:
            continue

        if w in common_strings.known_file_extenstions_list:
            continue
        if w in common_strings.known_toplevel_domains_list:
            continue
        if w in common_strings.country_toplevel_domains_ccTLD:
            continue
        if w in common_strings.private_toplevel_domains:
            continue
        if w in common_strings.private_toplevel_domains_IDN:
            continue

        new_words.append(w)

        # be sure that there are no single quotes in the string before priniting
        if "'" in w:
            print(f'    "{w}",')
        else:
            print(f"    '{w}',")

    print(f"{len(new_words)=}")
    # print entries for new words
    if len(new_words) > 0:
        # process acronyms and output English and Swedish for entries that are not jargon or 'nät- och sms-språk'
        for index, row in enumerate(ws.rows, start=2):
            e_word = ws.cell(index, eng_column)
            s_word = ws.cell(index, sv_column)
            a_word = ws.cell(index, acronym_column)

            ctg = ws.cell(index, category_column)
            # skip integer values
            if ctg.value is not None and isinstance(ctg.value, int):
                continue

            if isinstance(ctg.value, str):
                if ctg.value and ctg.value[0] == '=': # skip formulas
                    continue
                # if 'jargon' in ctg.value:
                #     continue
                # if 'nät- och sms-språk' in ctg.value:
                #     continue

            if isinstance(e_word.value, str):
                ewv=e_word.value.strip()
                ewv=ewv.replace('\j', ' ')

                if ewv == 'spamment':
                    print(f"FOUND: '{ewv=}'")

            if e_word.value not in new_words:
                continue

            output_line='['

            if isinstance(a_word.value, str):
                output_line=output_line + f"'{a_word.value}', "
            if isinstance(s_word.value, str):
                swv=s_word.value.strip()
                swv=swv.replace('\j', ' ')
                if swv == ewv:
                    output_line=output_line + f"'{ewv}'" + ","
                else:
                    output_line=output_line + f"'{ewv}'" + ", {'swe': '" + f"{swv}" + "'}],"
            else:
                output_line=output_line + f"'{ewv}'" + "],"

            if isinstance(ctg.value, str) and len(ctg.value) > 0:
                output_line=output_line + ' # ' + f"{ctg.value}"
            print(f"NEW: {output_line}")        


    # look for words where the Swedish term is the same ans the word
    if False:
        for w in common_acronyms.less_well_known_acronyms_list:
            if len(w) < 3:
                continue
            if w[1] and w[2]:
                if w[1] == w[2].get('swe'):
                    print(f"{w=}")

    # IT_ord_persons
    new_names=set()
    for fc in filter_categories:
        if fc == 'person':
            for p in IT_ord_persons:
                cs=p.count(',')
                if cs > 1:
                    print(f"More than one comma: {p=}")
                    if p == 'Icaza, de, Miguel' or p == 'Neumann, von, John':
                        ps=p.split(',')
                elif cs == 1:
                    ps=p.split(',')
                else:
                    ps=[p]
                for pp in ps:
                    pp=pp.strip()
                    if pp not in common_english.names_of_persons:
                        new_names.add(pp)

        new_names=sorted(new_names)
        print(f"{new_names=}")



    if not options.acronym_option:
        return

    # process acronyms and output English and Swedish for entries that are not jargon or 'nät- och sms-språk'
    for index, row in enumerate(ws.rows, start=2):
        a_word = ws.cell(index, acronym_column)
        e_word = ws.cell(index, eng_column)
        s_word = ws.cell(index, sv_column)

        ctg = ws.cell(index, category_column)
        # skip integer values
        if ctg.value is not None and isinstance(ctg.value, int):
            continue

        if isinstance(ctg.value, str):
            if ctg.value and ctg.value[0] == '=': # skip formulas
                continue
            if 'jargon' in ctg.value:
                continue
            if 'nät- och sms-språk' in ctg.value:
                continue
        if isinstance(a_word.value, str):
            awv=a_word.value.strip()

            if isinstance(e_word.value, str):
                ewv=e_word.value.strip()
                ewv=ewv.replace('\j', ' ')

            output_line='['
            if isinstance(s_word.value, str):
                swv=s_word.value.strip()
                swv=swv.replace('\j', ' ')
                output_line=output_line + f"'{awv}'" + ', ' + F"'{ewv}'" + ", {'swe': '" + f"{swv}" + "'}],"
            else:
                output_line=output_line + f"'{awv}'" + ', ' + F"'{ewv}'" + "],"

            if isinstance(ctg.value, str) and len(ctg.value) > 0:
                output_line=output_line + ' # ' + f"{ctg.value}"

            possible=False
            a_match=False
            for w in common_acronyms.well_known_acronyms_list:
                if len(w) < 2:
                    print(f"Too short error in entry: {w}")
                    continue
                if awv == w[0]:
                    possible=True
                if ewv == w[1] or  ewv == w[1]:
                    a_match=True

            if a_match:
                print(f"KNOWN: {output_line}")
            else:
                if possible:
                    print(f"POSSIBLE: {output_line}")
                else:
                    print(f"NEW: {output_line}")

            possible=False
            a_match=False
            for w in common_acronyms.less_well_known_acronyms_list:
                if len(w) < 2:
                    print(f"Too short error in entry: {w}")
                    continue
                if awv == w[0]:
                    possible=True
                if ewv == w[1] or  ewv == w[1]:
                    a_match=True

            if a_match:
                print(f"KNOWN: {output_line}")
            else:
                if possible:
                    print(f"POSSIBLE: {output_line}")
                else:
                    print(f"NEW: {output_line}")

#
if __name__ == "__main__": main()

